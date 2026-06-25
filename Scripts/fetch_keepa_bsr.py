"""
Fetch daily BSR data from the Keepa API and write one CSV per brand into
data/BSR/<brand>/<YY-MM-DD>.csv -- the SAME format the manual Keepa exports
use, so build_bsr_snapshots.py reads them with zero changes.

ASIN list + brand mapping come from data/master/sku_master.xlsx.
Domain: India ('IN').  Each ASIN with stats = 1 token.
The keepa library auto-waits for tokens, so big runs just take longer.

One-time setup:
    pip install keepa openpyxl
    setx KEEPA_API_KEY "your_new_private_key"   # then REOPEN the terminal

Run:
    python Scripts/fetch_keepa_bsr.py
"""
import csv, os, sys
from datetime import datetime
from pathlib import Path

try:
    import keepa
except ImportError:
    sys.exit("X  keepa not installed. Run:  pip install keepa")
try:
    import openpyxl
except ImportError:
    sys.exit("X  openpyxl not installed. Run:  pip install openpyxl")

SCRIPT_DIR   = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
MASTER_XLSX  = PROJECT_ROOT / "data" / "master" / "sku_master.xlsx"
BSR_DIR      = PROJECT_ROOT / "data" / "BSR"
DOMAIN       = "IN"

WANTED_BRANDS = {"Nexlev", "Audio Array", "Tonor", "White Mulberry"}

# Keepa CsvType indices (pulled from the library's own table so they stay
# correct even if a future version reorders them).
_IDX = {name: i for i, name, *_ in keepa.csv_indices}
SALES         = _IDX["SALES"]          # 3
RATING        = _IDX["RATING"]         # 16
COUNT_REVIEWS = _IDX["COUNT_REVIEWS"]  # 17

CSV_HEADER = [
    "Image", "Title",
    "Sales Rank: Current", "Sales Rank: 30 days avg.",
    "Sales Rank: 90 days avg.", "Sales Rank: 365 days avg.",
    "Sales Rank: Drops last 30 days",
    "Reviews: Rating", "Reviews: Rating Count",
    "ASIN", "Brand",
    "Monthly Sales Trends: Monthly Sold (Last Known)",
]

def load_asins():
    wb = openpyxl.load_workbook(MASTER_XLSX, read_only=True)
    ws = wb["Sheet1"]
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    ai, bi = hdr.index("ASIN"), hdr.index("Brand")
    by_brand, seen = {}, set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        asin, brand = r[ai], r[bi]
        if not asin or not brand or brand not in WANTED_BRANDS:
            continue
        asin = str(asin).strip()
        if len(asin) != 10 or (brand, asin) in seen:
            continue
        seen.add((brand, asin))
        by_brand.setdefault(brand, []).append(asin)
    return by_brand

def idx(arr, i):
    try:
        v = arr[i]
        return None if v in (None, -1) else v
    except (IndexError, TypeError):
        return None

def first_image(p):
    imgs = p.get("imagesCSV") or ""
    if imgs:
        f = imgs.split(",")[0].strip()
        if f:
            return f"https://m.media-amazon.com/images/I/{f}"
    return ""

def cell(v):
    return "" if v is None else v

CHUNK_SIZE  = 100   # ASINs per API call; a network drop loses at most this many
MAX_RETRIES = 5     # retries per chunk on a network/connection error

def row_for(p, brand):
    s   = p.get("stats") or {}
    cur = s.get("current") or []
    rating_raw = idx(cur, RATING)
    rating = round(rating_raw / 10, 1) if rating_raw is not None else None
    ms = p.get("monthlySold")
    if ms in (None, -1):
        ms = ""
    return [
        first_image(p),
        p.get("title", "") or "",
        cell(idx(cur, SALES)),
        cell(idx(s.get("avg30")  or [], SALES)),
        cell(idx(s.get("avg90")  or [], SALES)),
        cell(idx(s.get("avg365") or [], SALES)),
        cell(s.get("salesRankDrops30")),
        cell(rating),
        cell(idx(cur, COUNT_REVIEWS)),
        p.get("asin", ""),
        brand,
        ms,
    ]

def query_with_retry(api, chunk):
    """Query one chunk, retrying on transient network errors (no internet,
    DNS failure, dropped connection) so a brief blip doesn't kill the run."""
    import time
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            return api.query(chunk, domain=DOMAIN, stats=1, history=False, wait=True)
        except Exception as e:
            transient = any(t in str(e).lower() for t in
                            ("getaddrinfo", "connection", "timed out", "timeout",
                             "max retries", "resolve", "temporarily"))
            if transient and attempt < MAX_RETRIES:
                wait = 30 * attempt
                print(f"      network error (attempt {attempt}/{MAX_RETRIES}): "
                      f"retrying in {wait}s ...")
                time.sleep(wait)
                continue
            raise

def main():
    api_key = os.environ.get("KEEPA_API_KEY")
    if not api_key:
        sys.exit('X  KEEPA_API_KEY not set. Run:  setx KEEPA_API_KEY "your_key"  then reopen terminal')
    if not MASTER_XLSX.exists():
        sys.exit(f"X  Master file not found: {MASTER_XLSX}")

    by_brand = load_asins()
    total = sum(len(v) for v in by_brand.values())
    print(f"Loaded {total} ASINs across {len(by_brand)} brands:")
    for b, lst in by_brand.items():
        print(f"   {b:16s} {len(lst)}")

    print(f"\nConnecting to Keepa (domain={DOMAIN}) ...")
    api = keepa.Keepa(api_key)
    print(f"   tokens available now: {api.tokens_left}")

    datestamp = datetime.now().strftime("%y-%m-%d")

    for brand, asins in by_brand.items():
        out_dir = BSR_DIR / brand
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"{datestamp}.csv"

        # Resume support: if today's file already has all rows, skip this brand.
        if out_path.exists():
            try:
                with open(out_path, encoding="utf-8-sig") as f:
                    existing = sum(1 for _ in f) - 1   # minus header
            except OSError:
                existing = 0
            if existing >= len(asins):
                print(f"\n{brand}: already done today ({existing} rows) -- skipping.")
                continue
            else:
                print(f"\n{brand}: partial file found ({existing} rows) -- refetching.")

        print(f"\n{brand}: requesting {len(asins)} ASINs "
              f"in chunks of {CHUNK_SIZE} (auto-waits for tokens) ...")

        rows = []
        for i in range(0, len(asins), CHUNK_SIZE):
            chunk = asins[i:i + CHUNK_SIZE]
            print(f"   chunk {i//CHUNK_SIZE + 1}: {len(chunk)} ASINs ...")
            products = query_with_retry(api, chunk)
            for p in products:
                if p:
                    rows.append(row_for(p, brand))

        # Write only after the whole brand succeeded, so the file is never partial.
        with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(CSV_HEADER)
            w.writerows(rows)
        print(f"   wrote {out_path.relative_to(PROJECT_ROOT)}  ({len(rows)} rows)")

    print(f"\nDone. Next run: python Scripts/build_bsr_snapshots.py")

if __name__ == "__main__":
    main()
